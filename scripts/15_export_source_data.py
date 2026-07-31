# ============================================================
# 15_EXPORT_SOURCE_DATA.PY
#
# Purpose:
#   Build an Excel workbook containing the source data behind
#   every thesis figure — one tab per figure, tabs named
#   "Figure 1" through "Figure 8". Each tab shows:
#     * a title line
#     * a short caption
#     * the figure image itself (embedded)
#     * the numeric source data with clear column labels
#
# Output:
#   outputs/15_source_data/source_data_Amanli.xlsx
# ============================================================

import os
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from scipy.stats import binomtest

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")

OUT_DIR = os.path.join(OUTPUT_DIR, "15_source_data")
os.makedirs(OUT_DIR, exist_ok=True)
OUT_XLSX = os.path.join(OUT_DIR, "source_data_Amanli.xlsx")

# GitHub repository (public) — full URL used per figure and in the README
GITHUB_BASE = ("https://github.com/ZahraAmanli02/thesis_lfp_analysis_"
               "/blob/main/scripts")

# per-figure code URLs (both the figure-drawing script and the analysis script
# that produced its underlying numbers)
FIG_CODE_URLS = {
    1: {"figure": f"{GITHUB_BASE}/10b5_bootstrap_heatmap_clean.py",
        "analysis": f"{GITHUB_BASE}/10b_bootstrap_full.py"},
    2: {"figure": f"{GITHUB_BASE}/10b3_bootstrap.py",
        "analysis": f"{GITHUB_BASE}/10b_bootstrap_full.py"},
    3: {"figure": f"{GITHUB_BASE}/10b4_bootstrap.py",
        "analysis": f"{GITHUB_BASE}/10b_bootstrap_full.py"},
    4: {"figure": f"{GITHUB_BASE}/11c2_bootstrap_weight_top10.py",
        "analysis": f"{GITHUB_BASE}/11c1_bootstrap_weight_full.py"},
    5: {"figure": f"{GITHUB_BASE}/11c3_bootstrap_weight_per_phase.py",
        "analysis": f"{GITHUB_BASE}/11c1_bootstrap_weight_full.py"},
    6: {"figure": f"{GITHUB_BASE}/11c5_within_mouse_correlation.py",
        "analysis": f"{GITHUB_BASE}/11c5_within_mouse_correlation.py"},
    7: {"figure": f"{GITHUB_BASE}/12_rq3_cross_task_bootstrap.py",
        "analysis": f"{GITHUB_BASE}/12_rq3_cross_task_bootstrap.py"},
    8: {"figure": f"{GITHUB_BASE}/12b_rq3_band_vs_ratio_bootstrap.py",
        "analysis": f"{GITHUB_BASE}/12b_rq3_band_vs_ratio_bootstrap.py"},
}

# pooled feature CSVs (input to the models — the raw source data)
CABLE1_FEAT_CSV = os.path.join(
    OUTPUT_DIR, "10a_features_Cable1", "10a_features_Cable1.csv"
)
CABLE3_FEAT_CSV = (
    "/Users/amanlizahra/Desktop/For CABLE 3/thesis_lfp_analysis/"
    "outputs/10a_features_Cable3/10a_features_Cable3.csv"
)

# per-iteration bootstrap CSVs (numbers that were averaged into the summary)
RQ1_ITER_CSV = os.path.join(
    OUTPUT_DIR, "10b_bootstrap_full", "10b_bootstrap_iterations.csv"
)
RQ2_ITER_CSV = os.path.join(
    OUTPUT_DIR, "11c1_bootstrap_weight_full",
    "11c1_bootstrap_weight_iterations.csv"
)

BANDS = ["delta", "theta", "beta", "low_gamma", "high_gamma", "fast_gamma"]
SLOW_BANDS = ["delta", "theta", "beta"]
GAMMA_BANDS = ["low_gamma", "high_gamma", "fast_gamma"]
RATIOS = [
    "theta_delta", "beta_delta", "low_gamma_delta",
    "high_gamma_delta", "fast_gamma_delta",
    "beta_theta", "low_gamma_theta", "high_gamma_theta", "fast_gamma_theta",
    "low_gamma_beta", "high_gamma_beta", "fast_gamma_beta",
    "high_gamma_low_gamma", "fast_gamma_low_gamma", "fast_gamma_high_gamma",
]
CELLS = BANDS + RATIOS
ESTROUS_PHASES = ["A", "B", "C", "D"]
PHASE_NAMES = {"A": "pro-estrus", "B": "estrus",
               "C": "metestrus", "D": "diestrus"}

RQ1_CSV = os.path.join(
    OUTPUT_DIR, "10b_bootstrap_full", "10b_bootstrap_results_long.csv"
)
RQ2_CSV = os.path.join(
    OUTPUT_DIR, "11c1_bootstrap_weight_full",
    "11c1_bootstrap_weight_results_long.csv"
)
WITHIN_CSV = os.path.join(
    OUTPUT_DIR, "11c5_within_mouse_correlation",
    "11c5_within_mouse_correlations.csv"
)

# --- figure PNGs to embed ---
FIG_PNG = {
    1: os.path.join(OUTPUT_DIR, "10b5_bootstrap_heatmap_clean",
                    "10b5_bootstrap_heatmap_clean.png"),
    2: os.path.join(OUTPUT_DIR, "10b3_bootstrap", "10b3_top10_bar.png"),
    3: os.path.join(OUTPUT_DIR, "10b4_bootstrap", "10b4_per_phase_svm.png"),
    4: os.path.join(OUTPUT_DIR, "11c2_bootstrap_weight_top10",
                    "11c2_top10_bar.png"),
    5: os.path.join(OUTPUT_DIR, "11c3_bootstrap_weight_per_phase",
                    "11c3_per_phase_hfd_vs_ctrl.png"),
    6: os.path.join(OUTPUT_DIR, "11c5_within_mouse_correlation",
                    "11c5_within_mouse_bar.png"),
    7: os.path.join(OUTPUT_DIR, "12_rq3_cross_task_bootstrap",
                    "12_cross_task_scatter.png"),
    8: os.path.join(OUTPUT_DIR, "12b_rq3_band_vs_ratio_bootstrap",
                    "12b_band_vs_ratio.png"),
}

FIG_META = {
    1: {
        "title": "Figure 1 — RQ1 bootstrap heatmap (SVM + Random Forest)",
        "caption": ("Bootstrap mean balanced accuracy for each (estrous "
                    "phase × feature) combination, computed by pooling "
                    "Cable 1 + Cable 3 and running 1000 mouse-cluster "
                    "bootstrap iterations per cell. Two classifiers shown: "
                    "SVM-RBF (primary) and Random Forest (robustness check). "
                    "Chance = 0.500."),
        "result": ("Elevated accuracies (>0.60) concentrate in two zones: "
                   "pro-estrus (Phase A) × gamma cross-band ratios, and "
                   "estrus (Phase B) × delta-family cells. Metestrus (C) "
                   "and diestrus (D) sit near chance. Both classifiers "
                   "produce a similar pattern."),
    },
    2: {
        "title": "Figure 2 — RQ1 top-10 (SVM, ranked by 95% CI lower bound)",
        "caption": ("Ten (phase × feature) combinations ranked by the 95% CI "
                    "lower bound (most confident above chance first). "
                    "Highlighted bar = 95% CI excludes chance."),
        "result": ("The only cell whose full 95% CI clears chance is "
                   "Phase B × fast_gamma / delta (mean 0.68, CI [0.51, 0.86]). "
                   "The remaining top-10 cells reach mean accuracies of "
                   "0.61–0.72 but their CIs cross 0.5, so they are trending "
                   "rather than statistically confirmed."),
    },
    3: {
        "title": "Figure 3 — RQ1 per-phase overview (SVM)",
        "caption": ("Bootstrap mean balanced accuracy across all 21 feature "
                    "cells (6 bands + 15 ratios) per estrous phase. Highlighted "
                    "bar = 95% CI excludes chance."),
        "result": ("Estrus (B) shows a convergent delta-family pattern: "
                   "delta itself (0.63) and all /delta ratios (0.62–0.68) "
                   "elevated together. Pro-estrus (A) shows isolated high "
                   "gamma cross-band ratios. Metestrus (C) and diestrus (D) "
                   "hover around chance."),
    },
    4: {
        "title": "Figure 4 — RQ2 top-10 HFD body-weight regression (R²)",
        "caption": ("Ten (phase × feature) combinations for HFD-only body-"
                    "weight regression, ranked by bootstrap mean R². Target = "
                    "absolute body weight. All values negative — no cross-mouse "
                    "combination outperforms a 'predict the mean' baseline."),
        "result": ("Every one of the 84 cross-mouse (phase × feature) "
                   "combinations produced negative R². Cross-mouse regression "
                   "could not recover the weight signal from single-cell LFP "
                   "features — between-mouse baseline weight variance dominates."),
    },
    5: {
        "title": "Figure 5 — RQ2 per-phase overview, HFD vs CTRL",
        "caption": ("Per-phase body-weight regression R² for HFD (left) and "
                    "CTRL (right) mice across all 21 features. Both subsets "
                    "produce strongly negative R² in every phase — diagnostic "
                    "of a between-mouse baseline-variance ceiling."),
        "result": ("The HFD-vs-CTRL symmetry (both panels equally negative) "
                   "is diagnostic: the failure is not driven by the diet "
                   "effect itself but by the dominance of between-mouse "
                   "baseline weight variance (25–38 g). Cross-mouse "
                   "regression is not feasible in this design."),
    },
    6: {
        "title": "Figure 6 — Within-mouse LFP–weight correlation (HFD)",
        "caption": ("For each of 9 HFD mice with ≥4 diet recordings, Spearman "
                    "correlation between each LFP feature and the mouse's own "
                    "weight_delta was computed. Bars show median across mice, "
                    "dots show individual mice. Highlighted bars = sign-test "
                    "p < 0.05 (majority same-sign)."),
        "result": ("Four features reach significance and all four are "
                   "gamma / beta ratios. Strongest: fast_gamma / beta with "
                   "9 out of 9 HFD mice positive (median r = +0.37, sign-test "
                   "p = 0.004). The gamma-to-beta balance tracks weight gain "
                   "at the individual-animal level — a signal the cross-mouse "
                   "regression could not recover."),
    },
    7: {
        "title": "Figure 7 — Cross-task consistency (RQ1 vs RQ2)",
        "caption": ("Each dot is one (phase × feature) combination, plotted by "
                    "its RQ1 balanced accuracy (x-axis) and RQ2 R² for the HFD "
                    "body-weight regression (y-axis). A dissociation between "
                    "the two tasks is quantified by the Pearson correlation "
                    "(see top of this sheet)."),
        "result": ("The two tasks do not share a common set of informative "
                   "features. No cell rises above both chance lines "
                   "simultaneously, and the point-estimate correlation "
                   "across cells is not positive. Diet classification and "
                   "weight regression rely on distinct LFP substrates."),
    },
    8: {
        "title": "Figure 8 — Band vs ratio + slow vs gamma comparison",
        "caption": ("Distribution of SVM balanced accuracy per phase, split "
                    "either by feature type (single-band vs cross-frequency "
                    "ratio) or, within bands, by frequency range (slow = δθβ, "
                    "gamma = low/high/fast γ). Bars/dots show individual cells; "
                    "horizontal marks show group mean."),
        "result": ("Feature-type informativeness is phase-dependent: in pro-"
                   "estrus, ratios beat single bands (0.53 vs 0.45), while in "
                   "estrus single bands beat ratios (0.56 vs 0.51) — delta "
                   "dominates. Overall, slow bands and gamma bands are "
                   "roughly comparable (0.50 vs 0.50), which does not support "
                   "a specifically gamma-driven diet signal."),
    },
}


def format_cell(cell):
    if cell in BANDS:
        return cell
    for b in sorted(BANDS, key=len, reverse=True):
        if cell.startswith(b + "_"):
            rest = cell[len(b) + 1:]
            if rest in BANDS:
                return f"{b} / {rest}"
    return cell


def band_speed(cell):
    if cell in SLOW_BANDS:
        return "slow"
    if cell in GAMMA_BANDS:
        return "gamma"
    return ""


# ============================================================
# LOAD SOURCE CSVs
# ============================================================

df_rq1 = pd.read_csv(RQ1_CSV)
df_rq2 = pd.read_csv(RQ2_CSV)
df_within = pd.read_csv(WITHIN_CSV)

print(f"RQ1 rows loaded: {len(df_rq1)}")
print(f"RQ2 rows loaded: {len(df_rq2)}")
print(f"Within-mouse rows loaded: {len(df_within)}")


# ============================================================
# BUILD PER-SHEET DATAFRAMES
# ============================================================

def add_phase_name(df):
    df = df.copy()
    df.insert(df.columns.get_loc("phase") + 1, "phase_name",
              df["phase"].map(PHASE_NAMES))
    return df


def add_feature_pretty(df, cell_col="cell"):
    df = df.copy()
    df.insert(df.columns.get_loc(cell_col) + 1, "feature_cell_pretty",
              df[cell_col].apply(format_cell))
    return df


# --- Figure 1: RQ1 heatmap — SVM and RF side by side (wide format) ---
def _rq1_wide():
    base_cols = ["phase", "cell", "cell_type", "n_recordings", "n_mice"]
    # take the first row per (phase, cell) — n_recordings / n_mice identical
    # across models by construction
    base = (df_rq1[base_cols]
            .drop_duplicates(subset=["phase", "cell"])
            .reset_index(drop=True))
    svm_side = (df_rq1[df_rq1["model"] == "svm_rbf"]
                [["phase", "cell", "boot_mean", "boot_ci_lo", "boot_ci_hi"]]
                .rename(columns={
                    "boot_mean": "SVM_mean_balanced_accuracy",
                    "boot_ci_lo": "SVM_ci_95_lower",
                    "boot_ci_hi": "SVM_ci_95_upper",
                }))
    rf_side = (df_rq1[df_rq1["model"] == "random_forest"]
               [["phase", "cell", "boot_mean", "boot_ci_lo", "boot_ci_hi"]]
               .rename(columns={
                   "boot_mean": "RF_mean_balanced_accuracy",
                   "boot_ci_lo": "RF_ci_95_lower",
                   "boot_ci_hi": "RF_ci_95_upper",
               }))
    wide = (base.merge(svm_side, on=["phase", "cell"], how="left")
                .merge(rf_side, on=["phase", "cell"], how="left"))
    wide = wide.rename(columns={"cell": "feature_cell"})
    return wide


fig1 = _rq1_wide()
fig1 = add_phase_name(fig1)
fig1 = add_feature_pretty(fig1, "feature_cell")
fig1["cell_order"] = fig1["feature_cell"].apply(
    lambda c: CELLS.index(c) if c in CELLS else 999)
fig1 = (fig1.sort_values(["phase", "cell_order"])
             .drop(columns=["cell_order"])
             .reset_index(drop=True))
# column order: identifiers first, SVM block, RF block
_ordered = (["phase", "phase_name", "feature_cell", "feature_cell_pretty",
             "cell_type", "n_recordings", "n_mice",
             "SVM_mean_balanced_accuracy", "SVM_ci_95_lower", "SVM_ci_95_upper",
             "RF_mean_balanced_accuracy", "RF_ci_95_lower", "RF_ci_95_upper"])
fig1 = fig1[[c for c in _ordered if c in fig1.columns]]


# --- Figure 2: RQ1 top-10 SVM by CI lower bound ---
svm_all = df_rq1[df_rq1["model"] == "svm_rbf"].dropna(subset=["boot_mean"]).copy()
top10 = (svm_all
         .sort_values(["boot_ci_lo", "boot_mean"], ascending=[False, False])
         .head(10).reset_index(drop=True))
fig2 = pd.DataFrame({
    "rank": range(1, len(top10) + 1),
    "phase": top10["phase"],
    "phase_name": top10["phase"].map(PHASE_NAMES),
    "feature_cell": top10["cell"],
    "feature_cell_pretty": top10["cell"].apply(format_cell),
    "cell_type": top10["cell_type"],
    "mean_balanced_accuracy": top10["boot_mean"].round(3),
    "ci_95_lower": top10["boot_ci_lo"].round(3),
    "ci_95_upper": top10["boot_ci_hi"].round(3),
    "ci_confirmed_above_chance": top10["boot_ci_lo"] > 0.5,
    "n_recordings": top10["n_recordings"],
    "n_mice": top10["n_mice"],
})


# --- Figure 3: RQ1 per-phase overview SVM ---
fig3 = (svm_all[["phase", "cell", "cell_type",
                 "boot_mean", "boot_ci_lo", "boot_ci_hi",
                 "n_recordings", "n_mice"]]
        .rename(columns={"cell": "feature_cell",
                         "boot_mean": "mean_balanced_accuracy",
                         "boot_ci_lo": "ci_95_lower",
                         "boot_ci_hi": "ci_95_upper"}))
fig3["ci_confirmed_above_chance"] = fig3["ci_95_lower"] > 0.5
fig3 = add_phase_name(fig3)
fig3 = add_feature_pretty(fig3, "feature_cell")
fig3["cell_order"] = fig3["feature_cell"].apply(
    lambda c: CELLS.index(c) if c in CELLS else 999)
fig3 = (fig3.sort_values(["phase", "cell_order"])
             .drop(columns=["cell_order"]).reset_index(drop=True))


# --- Figure 4: RQ2 HFD top-10 R² ---
hfd_all = df_rq2[df_rq2["subset"] == "HFD"].dropna(subset=["boot_mean"]).copy()
top10_hfd = (hfd_all.sort_values("boot_mean", ascending=False)
             .head(10).reset_index(drop=True))
fig4 = pd.DataFrame({
    "rank": range(1, len(top10_hfd) + 1),
    "phase": top10_hfd["phase"],
    "phase_name": top10_hfd["phase"].map(PHASE_NAMES),
    "feature_cell": top10_hfd["cell"],
    "feature_cell_pretty": top10_hfd["cell"].apply(format_cell),
    "cell_type": top10_hfd["cell_type"],
    "mean_R2": top10_hfd["boot_mean"].round(3),
    "ci_95_lower": top10_hfd["boot_ci_lo"].round(3),
    "ci_95_upper": top10_hfd["boot_ci_hi"].round(3),
    "n_recordings": top10_hfd["n_recordings"],
    "n_mice": top10_hfd["n_mice"],
    "target": "body_weight (absolute)",
})


# --- Figure 5: RQ2 HFD vs CTRL per phase ---
# Include the `note` column so rows with empty R²/CI are self-explanatory
# (e.g. "phase_skipped_too_few_mice" for HFD × Phase C in this dataset).
fig5 = (df_rq2[["subset", "phase", "cell", "cell_type",
                "boot_mean", "boot_ci_lo", "boot_ci_hi",
                "n_recordings", "n_mice", "note"]]
        .rename(columns={"cell": "feature_cell",
                         "boot_mean": "mean_R2",
                         "boot_ci_lo": "ci_95_lower",
                         "boot_ci_hi": "ci_95_upper",
                         "note": "status_note"}))
# Fill in a friendly status when the row has no computed values
fig5["status_note"] = fig5.apply(
    lambda r: (r["status_note"] if isinstance(r["status_note"], str)
               and r["status_note"] != "" else "ok")
    if pd.notna(r["mean_R2"])
    else (r["status_note"] if isinstance(r["status_note"], str)
          and r["status_note"] != ""
          else "no model fit (eligibility criterion not met)"),
    axis=1,
)
fig5 = add_phase_name(fig5)
fig5 = add_feature_pretty(fig5, "feature_cell")
fig5["cell_order"] = fig5["feature_cell"].apply(
    lambda c: CELLS.index(c) if c in CELLS else 999)
fig5 = (fig5.sort_values(["subset", "phase", "cell_order"])
             .drop(columns=["cell_order"]).reset_index(drop=True))


# --- Figure 6: within-mouse correlations, two sub-tables ---
fig6_per_mouse = df_within.rename(columns={
    "mouse_uid": "mouse",
    "n_rec": "n_recordings_this_mouse",
    "spearman_r": "spearman_correlation",
    "spearman_p": "spearman_p_value",
})[["mouse", "feature", "n_recordings_this_mouse",
    "spearman_correlation", "spearman_p_value"]]


def summarize_feature(sub):
    r_vals = sub["spearman_correlation"].to_numpy()
    n_pos = int((r_vals > 0).sum())
    n_neg = int((r_vals < 0).sum())
    n_nonzero = n_pos + n_neg
    if n_nonzero == 0:
        sign_p = 1.0
    else:
        sign_p = float(binomtest(max(n_pos, n_neg), n_nonzero,
                                 p=0.5, alternative="two-sided").pvalue)
    return pd.Series({
        "n_mice": len(sub),
        "median_r": float(np.median(r_vals)),
        "mean_r": float(np.mean(r_vals)),
        "n_positive": n_pos,
        "n_negative": n_neg,
        "sign_test_p": sign_p,
        "significant_at_0.05": sign_p < 0.05,
    })


fig6_summary = (fig6_per_mouse.groupby("feature", as_index=False)
                              .apply(summarize_feature, include_groups=False)
                              .sort_values("median_r", ascending=False)
                              .reset_index(drop=True))


# --- Figure 7: cross-task scatter data ---
svm_ct = (df_rq1[df_rq1["model"] == "svm_rbf"].dropna(subset=["boot_mean"])
          [["phase", "cell", "cell_type",
            "boot_mean", "boot_ci_lo", "boot_ci_hi"]]
          .rename(columns={"cell": "feature_cell",
                           "boot_mean": "RQ1_mean_balanced_accuracy",
                           "boot_ci_lo": "RQ1_ci_lower",
                           "boot_ci_hi": "RQ1_ci_upper"}))
r2_ct = (df_rq2[df_rq2["subset"] == "HFD"].dropna(subset=["boot_mean"])
         [["phase", "cell", "boot_mean", "boot_ci_lo", "boot_ci_hi"]]
         .rename(columns={"cell": "feature_cell",
                          "boot_mean": "RQ2_mean_R2",
                          "boot_ci_lo": "RQ2_ci_lower",
                          "boot_ci_hi": "RQ2_ci_upper"}))
fig7 = svm_ct.merge(r2_ct, on=["phase", "feature_cell"], how="inner")
fig7 = add_phase_name(fig7)
fig7 = add_feature_pretty(fig7, "feature_cell")
fig7 = fig7.sort_values(["phase", "feature_cell"]).reset_index(drop=True)

x_ct = fig7["RQ1_mean_balanced_accuracy"].to_numpy()
y_ct = fig7["RQ2_mean_R2"].to_numpy()
pearson_r = float(np.corrcoef(x_ct, y_ct)[0, 1])
n_above_both = int(((x_ct > 0.5) & (y_ct > 0.0)).sum())


# --- Figure 8: band vs ratio + slow vs gamma ---
svm_bv = df_rq1[df_rq1["model"] == "svm_rbf"].dropna(subset=["boot_mean"]).copy()
svm_bv["band_speed"] = svm_bv["cell"].apply(band_speed)
fig8_per_cell = (svm_bv[["phase", "cell", "cell_type", "band_speed",
                         "boot_mean"]]
                 .rename(columns={"cell": "feature_cell",
                                  "boot_mean": "mean_balanced_accuracy"}))
fig8_per_cell = add_phase_name(fig8_per_cell)
fig8_per_cell = add_feature_pretty(fig8_per_cell, "feature_cell")
fig8_per_cell = fig8_per_cell.sort_values(
    ["phase", "cell_type", "feature_cell"]).reset_index(drop=True)

records = []
for phase in ESTROUS_PHASES:
    for gtype in ("band", "ratio"):
        vals = svm_bv[(svm_bv["phase"] == phase)
                      & (svm_bv["cell_type"] == gtype)]["boot_mean"]
        if len(vals):
            records.append({"comparison": "band_vs_ratio",
                            "phase": phase,
                            "phase_name": PHASE_NAMES[phase],
                            "group": gtype,
                            "n_cells": len(vals),
                            "mean": float(vals.mean()),
                            "median": float(vals.median()),
                            "best": float(vals.max())})
    for gtype in ("slow", "gamma"):
        vals = svm_bv[(svm_bv["phase"] == phase)
                      & (svm_bv["band_speed"] == gtype)]["boot_mean"]
        if len(vals):
            records.append({"comparison": "slow_vs_gamma",
                            "phase": phase,
                            "phase_name": PHASE_NAMES[phase],
                            "group": gtype,
                            "n_cells": len(vals),
                            "mean": float(vals.mean()),
                            "median": float(vals.median()),
                            "best": float(vals.max())})
fig8_group = pd.DataFrame(records)


# ============================================================
# WRITE EXCEL — with title / caption / embedded image / table
# ============================================================

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
CAPTION_FONT = Font(name="Calibri", size=11, italic=True, color="404040")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="404040")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
THIN = Side(border_style="thin", color="B0B0B0")
CELL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

MAX_IMG_WIDTH_PX = 900          # scale images down so they fit next to the table
FIG_ROW_HEIGHT = 20             # approximate pixels per Excel row

# Human-readable description for every column that appears in any sheet.
COLUMN_DESCRIPTIONS = {
    "phase":                        "Estrous phase code (A/B/C/D)",
    "phase_name":                   "Estrous phase name (pro-estrus / estrus / metestrus / diestrus)",
    "subset":                       "HFD = high-fat diet mice; CTRL = control mice",
    "feature_cell":                 "LFP feature identifier (internal name, e.g. 'fast_gamma_delta')",
    "feature_cell_pretty":          "Human-readable feature name (e.g. 'fast_gamma / delta')",
    "cell_type":                    "'band' = single frequency band; 'ratio' = pairwise band-to-band ratio",
    "band_speed":                   "'slow' = δ / θ / β; 'gamma' = low / high / fast γ",
    "model":                        "Classifier used: 'svm_rbf' (SVM-RBF) or 'random_forest'",
    "n_recordings":                 "Number of recordings in this phase (or phase × subset)",
    "n_mice":                       "Number of unique mice contributing recordings",
    "mean_balanced_accuracy":       "Bootstrap mean of balanced accuracy over 1000 iterations (chance = 0.5)",
    "ci_95_lower":                  "Lower bound of the 95% bootstrap CI (2.5% percentile)",
    "ci_95_upper":                  "Upper bound of the 95% bootstrap CI (97.5% percentile)",
    "SVM_mean_balanced_accuracy":   "SVM-RBF bootstrap mean balanced accuracy",
    "SVM_ci_95_lower":              "SVM-RBF 95% CI lower bound",
    "SVM_ci_95_upper":              "SVM-RBF 95% CI upper bound",
    "RF_mean_balanced_accuracy":    "Random Forest bootstrap mean balanced accuracy",
    "RF_ci_95_lower":               "Random Forest 95% CI lower bound",
    "RF_ci_95_upper":               "Random Forest 95% CI upper bound",
    "rank":                         "Rank position (1 = strongest)",
    "ci_confirmed_above_chance":    "TRUE if the 95% CI lower bound is above chance",
    "mean_R2":                      "Bootstrap mean R² over 1000 iterations (chance = 0)",
    "target":                       "Target variable being predicted",
    "status_note":                  "Explanation when a row has no fitted values (e.g. eligibility criterion not met)",
    "mouse":                        "Mouse identifier",
    "feature":                      "LFP feature (internal name)",
    "n_recordings_this_mouse":      "Number of diet-phase recordings for this individual mouse",
    "spearman_correlation":         "Within-mouse Spearman r (feature vs weight_delta) across that mouse's recordings",
    "spearman_p_value":             "P-value of the within-mouse Spearman correlation",
    "median_r":                     "Median Spearman r across mice",
    "mean_r":                       "Mean Spearman r across mice",
    "n_positive":                   "Number of mice with a positive Spearman r",
    "n_negative":                   "Number of mice with a negative Spearman r",
    "sign_test_p":                  "Two-sided binomial sign-test p (null: r sign is 50/50 across mice)",
    "significant_at_0.05":          "TRUE if sign-test p < 0.05",
    "RQ1_mean_balanced_accuracy":   "RQ1 (diet classification) SVM bootstrap mean balanced accuracy",
    "RQ1_ci_lower":                 "RQ1 95% CI lower bound",
    "RQ1_ci_upper":                 "RQ1 95% CI upper bound",
    "RQ2_mean_R2":                  "RQ2 (HFD body-weight regression) bootstrap mean R²",
    "RQ2_ci_lower":                 "RQ2 95% CI lower bound",
    "RQ2_ci_upper":                 "RQ2 95% CI upper bound",
    "comparison":                   "Which comparison group: 'band_vs_ratio' or 'slow_vs_gamma'",
    "group":                        "Group label inside the comparison (band / ratio, or slow / gamma)",
    "n_cells":                      "Number of cells in this group",
    "mean":                         "Mean across cells in this group",
    "median":                       "Median across cells in this group",
    "best":                         "Maximum (best) value across cells in this group",
}


def scaled_image(path, max_width_px=MAX_IMG_WIDTH_PX):
    """Return an openpyxl Image scaled to fit max_width_px while keeping aspect."""
    with PILImage.open(path) as im:
        w, h = im.size
    scale = min(1.0, max_width_px / float(w))
    img = XLImage(path)
    img.width = int(w * scale)
    img.height = int(h * scale)
    return img, int(h * scale)


def write_dataframe_at(ws, df, start_row, start_col=1):
    """Write a DataFrame with a coloured header at the given position."""
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = CELL_BORDER
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            v = val
            if isinstance(v, (np.floating,)):
                v = float(v)
            elif isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.bool_,)):
                v = bool(v)
            elif pd.isna(v):
                v = None
            c = ws.cell(row=start_row + i, column=start_col + j, value=v)
            c.border = CELL_BORDER
            if isinstance(v, float):
                c.number_format = "0.000"
    # widen columns based on content length
    for j, col in enumerate(df.columns):
        max_len = max(
            [len(str(col))]
            + [len(str(x)) for x in df[col].astype(str).tolist()[:200]]
        )
        ws.column_dimensions[get_column_letter(start_col + j)].width = \
            min(max(10, max_len + 2), 40)


def write_figure_sheet(wb, fig_id, tables):
    """
    tables : list of tuples (section_title_or_None, DataFrame)
             None = no section header (single table)
    """
    ws_name = f"Figure {fig_id}"
    ws = wb.create_sheet(ws_name)
    meta = FIG_META[fig_id]

    # 1. Title
    ws.cell(row=1, column=1, value=meta["title"]).font = TITLE_FONT
    ws.row_dimensions[1].height = 22

    # 2. Caption (wrapped across a wide merged cell)
    ws.cell(row=3, column=1, value="What the figure shows:").font = SECTION_FONT
    ws.cell(row=4, column=1, value=meta["caption"]).font = CAPTION_FONT
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=12)
    ws.cell(row=4, column=1).alignment = Alignment(wrap_text=True,
                                                    vertical="top")
    ws.row_dimensions[4].height = 60

    # 3. Key result
    ws.cell(row=6, column=1, value="Key result:").font = SECTION_FONT
    result_cell = ws.cell(row=7, column=1, value=meta["result"])
    result_cell.font = Font(name="Calibri", size=11, color="1F4E79")
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=12)
    result_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[7].height = 70

    # 4. GitHub link — code that generated this figure and its numbers
    urls = FIG_CODE_URLS.get(fig_id, {})
    ws.cell(row=9, column=1, value="Code (GitHub):").font = SECTION_FONT
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    if urls.get("figure"):
        c = ws.cell(row=10, column=1,
                    value=f"Figure-drawing script: {urls['figure']}")
        c.font = link_font
        c.hyperlink = urls["figure"]
    if urls.get("analysis") and urls.get("analysis") != urls.get("figure"):
        c = ws.cell(row=11, column=1,
                    value=f"Analysis script (produces the numbers): "
                          f"{urls['analysis']}")
        c.font = link_font
        c.hyperlink = urls["analysis"]

    # 5. Embedded figure image
    img_row = 13
    img_path = FIG_PNG[fig_id]
    if os.path.exists(img_path):
        img, img_h_px = scaled_image(img_path)
        ws.add_image(img, f"A{img_row}")
        rows_taken = int(np.ceil(img_h_px / FIG_ROW_HEIGHT)) + 2
    else:
        ws.cell(row=img_row, column=1, value=f"[missing image: {img_path}]") \
            .font = CAPTION_FONT
        rows_taken = 3

    # 4. "Source data" section marker
    data_row = img_row + rows_taken + 1
    ws.cell(row=data_row, column=1, value="Source data").font = SECTION_FONT
    data_row += 2

    # 5. Column key — one row per column used across the tables in this sheet
    all_cols = []
    for _, df in tables:
        for c in df.columns:
            if c not in all_cols:
                all_cols.append(c)
    ws.cell(row=data_row, column=1, value="Column key").font = SECTION_FONT
    data_row += 1
    key_df = pd.DataFrame({
        "column": all_cols,
        "description": [COLUMN_DESCRIPTIONS.get(c, "") for c in all_cols],
    })
    write_dataframe_at(ws, key_df, start_row=data_row, start_col=1)
    data_row += len(key_df) + 3

    # 6. tables (one or many)
    for section_title, df in tables:
        if section_title:
            ws.cell(row=data_row, column=1, value=section_title).font = SECTION_FONT
            data_row += 1
        write_dataframe_at(ws, df, start_row=data_row, start_col=1)
        data_row += len(df) + 3   # gap before next table


def write_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.cell(row=1, column=1,
            value="Source data for the thesis figures — Amanli, MSc thesis") \
        .font = TITLE_FONT

    # GitHub link at the top of the README (clickable hyperlink)
    ws.cell(row=2, column=1, value="Code repository (GitHub):").font = SECTION_FONT
    github_url = "https://github.com/ZahraAmanli02/thesis_lfp_analysis_"
    link_cell = ws.cell(row=3, column=1, value=github_url)
    link_cell.font = Font(name="Calibri", size=11, color="0563C1",
                          underline="single")
    link_cell.hyperlink = github_url

    lines = [
        "",
        "Each sheet named 'Figure N' contains the numeric data behind the "
        "corresponding thesis figure, together with the figure image, a short "
        "caption, and a hyperlink to the code that generated the figure and "
        "its underlying numbers.",
        "",
        "For every figure sheet, the raw per-iteration bootstrap values that "
        "were averaged into the plotted means are included as a sub-table "
        "directly below the summary table on the same sheet. Figures 6 and 8 "
        "already show individual per-mouse / per-cell values in their own "
        "sub-tables and therefore do not need an additional iteration table.",
        "",
        "The workbook also contains one extra sheet with the model input data:",
        "  • 'Pooled features (input)' — the 208-recording feature table that "
        "was fed to every model (raw input source data, produced by 10a).",
        "",
        "Research questions covered:",
        "  • RQ1 — Can the LFP (CFD-derived band powers) classify diet group? "
        "That is, given a recording's LFP features, can a machine-learning "
        "model tell whether the animal is on the high-fat diet (HF) or the "
        "control diet (CTRL)? Metric: balanced accuracy (chance = 0.5). "
        "Figures 1, 2, 3 report RQ1.",
        "  • RQ2 — Can the LFP predict body weight? Specifically, can a "
        "regressor trained on LFP features predict how much weight the "
        "animal has gained from its own baseline (weight_delta)? Metric: "
        "R² (chance = 0). Figures 4, 5, 6 report RQ2 — separately for HFD "
        "and CTRL mice, and finally at the within-mouse level.",
        "  • RQ3 — How do the two tasks relate to each other? Do the same "
        "LFP features that classify diet also predict weight? Figures 7 "
        "and 8 report this cross-task view and a feature-type breakdown "
        "(bands vs ratios, slow vs gamma).",
        "",
        "Common conventions:",
        "  • estrous phase codes: A = pro-estrus, B = estrus, "
        "C = metestrus, D = diestrus",
        "  • subset: HFD = high-fat diet mice, CTRL = control mice",
        "  • cell_type: 'band' = single frequency band, "
        "'ratio' = pairwise band-to-band ratio",
        "  • band_speed: 'slow' = delta/theta/beta, "
        "'gamma' = low/high/fast gamma",
        "  • ci_95_lower / ci_95_upper: 2.5% and 97.5% percentiles of the "
        "1000-iteration bootstrap distribution (per phase × feature)",
        "  • feature_cell = internal name; feature_cell_pretty = human-readable "
        "(e.g. 'fast_gamma / delta')",
        "",
        "Abbreviations & terms:",
        "  • LFP = Local Field Potential (recorded electrophysiological signal)",
        "  • CFD = Channel-to-channel Differential — the input actually used "
        "for the models (difference between two recording channels; "
        "band powers are extracted from this signal)",
        "  • SVM-RBF = Support Vector Machine with a Radial Basis Function "
        "kernel; a supervised classifier that separates two groups (here "
        "HF vs CTRL) using a non-linear decision boundary in feature space. "
        "Used as the primary classifier.",
        "  • Random Forest = an ensemble of many decision trees that vote on "
        "the class label; used as an independent-algorithm sanity check "
        "alongside SVM-RBF.",
        "  • Random Forest regressor = the regression variant of Random "
        "Forest, used in RQ2 to predict weight_delta.",
        "  • balanced accuracy = average of sensitivity and specificity; "
        "unaffected by class imbalance. Chance level = 0.5.",
        "  • R² (coefficient of determination) = fraction of target variance "
        "explained by the model. Chance = 0 (i.e., 'always predict the "
        "mean'). Negative values mean the model does worse than that "
        "baseline.",
        "  • Spearman r = rank correlation coefficient; used in Figure 6 for "
        "within-mouse LFP-vs-weight correlations because it does not assume "
        "a linear relationship.",
        "  • bootstrap = repeatedly resample the data (here: sample mice with "
        "replacement) and re-fit the model to build up a distribution of "
        "the metric; the 2.5%–97.5% percentiles give the 95% CI.",
        "  • mouse-cluster bootstrap = resampling done at the animal level "
        "rather than at the recording level, so all recordings of a given "
        "mouse move together into either the training or the out-of-bag "
        "set — prevents the same animal from appearing in both sets.",
        "  • OOB (out-of-bag) = the mice that were NOT drawn in a given "
        "bootstrap iteration; used as the test set for that iteration.",
        "  • weight_delta = body_weight − that mouse's own baseline body "
        "weight (the mean of its diet_phase == 'baseline' recordings).",
        "",
        "Sheet list (each includes the figure image, its caption, its key "
        "result, and the underlying data):",
        "  • Figure 1  — RQ1 (diet classification): bootstrap heatmap "
        "(SVM + Random Forest)",
        "  • Figure 2  — RQ1: top-10, ranked by 95% CI lower bound",
        "  • Figure 3  — RQ1: per-phase overview (SVM)",
        "  • Figure 4  — RQ2 (weight prediction): top-10 HFD body-weight "
        "regression (R²)",
        "  • Figure 5  — RQ2: per-phase overview, HFD vs CTRL",
        "  • Figure 6  — RQ2: within-mouse LFP–weight correlation (HFD)",
        "                (two sub-tables: per-mouse values + per-feature summary)",
        "  • Figure 7  — RQ3 (cross-task view): RQ1 vs RQ2 per feature",
        "  • Figure 8  — RQ3: band vs ratio + slow vs gamma comparison",
        "                (two sub-tables: per-cell values + group summary)",
    ]
    for i, line in enumerate(lines, start=5):
        c = ws.cell(row=i, column=1, value=line)
        if line and not line.startswith("  "):
            c.font = SECTION_FONT
    ws.column_dimensions["A"].width = 100


# --- Build workbook ---
wb = Workbook()
# openpyxl gives us a default sheet — drop it, we build fresh
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

write_readme(wb)
# ============================================================
# LOAD PER-ITERATION CSVs (raw numbers averaged into each figure)
# ============================================================
rq1_iter = pd.read_csv(RQ1_ITER_CSV) if os.path.exists(RQ1_ITER_CSV) else None
rq2_iter = pd.read_csv(RQ2_ITER_CSV) if os.path.exists(RQ2_ITER_CSV) else None


def _iter_table(df, subset_col_map=None):
    """Format an iteration DataFrame for inclusion inside a figure sheet."""
    d = df.copy()
    if "cell" in d.columns:
        d.insert(d.columns.get_loc("cell") + 1, "cell_pretty",
                 d["cell"].apply(format_cell))
    if "phase" in d.columns:
        d.insert(d.columns.get_loc("phase") + 1, "phase_name",
                 d["phase"].map(PHASE_NAMES))
    return d


# ---- Figure 1: summary + all RQ1 iterations (SVM + RF, every cell) ----
fig1_iter_table = _iter_table(rq1_iter) if rq1_iter is not None else None
fig1_iter_hdr = (
    f"Raw source data — every bootstrap iteration that fed the averages "
    f"above ({0 if rq1_iter is None else len(rq1_iter):,} rows; "
    "84 phase×feature cells × 2 models × 1000 iterations)."
)

write_figure_sheet(wb, 1,
    [("Summary — bootstrap mean and 95% CI per phase × feature × model", fig1)]
    + ([(fig1_iter_hdr, fig1_iter_table)] if fig1_iter_table is not None else [])
)

# ---- Figure 2: summary + RQ1 iterations filtered to the top-10 cells ----
top2_keys = set(zip(fig2["phase"].tolist(), fig2["feature_cell"].tolist()))
if rq1_iter is not None:
    mask = ([(p, c) in top2_keys and m == "svm_rbf"
             for p, c, m in zip(rq1_iter["phase"], rq1_iter["cell"], rq1_iter["model"])])
    fig2_iter_table = _iter_table(rq1_iter[mask])
    fig2_iter_hdr = (
        f"Raw source data — every bootstrap iteration for the 10 (phase × "
        f"feature) cells shown above ({len(fig2_iter_table):,} rows; "
        "10 cells × SVM-RBF × 1000 iterations)."
    )
else:
    fig2_iter_table, fig2_iter_hdr = None, ""

write_figure_sheet(wb, 2,
    [("Summary — top 10 cells, mean bal_acc + 95% CI", fig2)]
    + ([(fig2_iter_hdr, fig2_iter_table)] if fig2_iter_table is not None else [])
)

# ---- Figure 3: summary + RQ1 iterations (SVM only, all cells) ----
if rq1_iter is not None:
    fig3_iter_table = _iter_table(rq1_iter[rq1_iter["model"] == "svm_rbf"])
    fig3_iter_hdr = (
        f"Raw source data — every bootstrap iteration for SVM-RBF, all cells "
        f"({len(fig3_iter_table):,} rows; 84 cells × 1 model × 1000 iterations)."
    )
else:
    fig3_iter_table, fig3_iter_hdr = None, ""

write_figure_sheet(wb, 3,
    [("Summary — per-phase overview, SVM-RBF", fig3)]
    + ([(fig3_iter_hdr, fig3_iter_table)] if fig3_iter_table is not None else [])
)

# ---- Figure 4: summary + RQ2 iterations filtered to top-10 HFD cells ----
top4_keys = set(zip(fig4["phase"].tolist(), fig4["feature_cell"].tolist()))
if rq2_iter is not None:
    mask = ([(p, c) in top4_keys and s == "HFD"
             for p, c, s in zip(rq2_iter["phase"], rq2_iter["cell"], rq2_iter["subset"])])
    fig4_iter_table = _iter_table(rq2_iter[mask])
    fig4_iter_hdr = (
        f"Raw source data — every bootstrap iteration for the 10 HFD (phase × "
        f"feature) cells shown above ({len(fig4_iter_table):,} rows)."
    )
else:
    fig4_iter_table, fig4_iter_hdr = None, ""

write_figure_sheet(wb, 4,
    [("Summary — top 10 HFD cells, mean R² + 95% CI", fig4)]
    + ([(fig4_iter_hdr, fig4_iter_table)] if fig4_iter_table is not None else [])
)

# ---- Figure 5: summary + RQ2 iterations (all HFD + CTRL cells) ----
if rq2_iter is not None:
    fig5_iter_table = _iter_table(rq2_iter)
    fig5_iter_hdr = (
        f"Raw source data — every bootstrap iteration for all HFD and CTRL "
        f"(phase × feature) cells ({len(fig5_iter_table):,} rows; "
        "84 cells × 2 subsets × 1000 iterations, less any skipped)."
    )
else:
    fig5_iter_table, fig5_iter_hdr = None, ""

write_figure_sheet(wb, 5,
    [("Summary — HFD vs CTRL per phase × feature (mean R² + 95% CI)", fig5)]
    + ([(fig5_iter_hdr, fig5_iter_table)] if fig5_iter_table is not None else [])
)
write_figure_sheet(wb, 6, [
    (f"Sub-table A — per-mouse × per-feature correlations "
     f"(n = {len(fig6_per_mouse)})", fig6_per_mouse),
    (f"Sub-table B — per-feature summary across mice "
     f"(median, sign test)", fig6_summary),
])
# Figure 7: summary + iterations from BOTH RQ1 and RQ2 (the two axes of the scatter)
fig7_header = (f"Summary — Pearson r between RQ1 balanced accuracy and RQ2 R² "
               f"across {len(fig7)} phase × feature cells: r = {pearson_r:+.3f}   "
               f"|   Cells with RQ1 > 0.5 AND RQ2 > 0: "
               f"{n_above_both} / {len(fig7)}")

fig7_iter_tables = []
if rq1_iter is not None:
    rq1_svm = rq1_iter[rq1_iter["model"] == "svm_rbf"]
    fig7_iter_tables.append((
        f"Raw source data (x-axis, RQ1) — every SVM-RBF bootstrap iteration "
        f"per phase × feature ({len(rq1_svm):,} rows).",
        _iter_table(rq1_svm),
    ))
if rq2_iter is not None:
    rq2_hfd = rq2_iter[rq2_iter["subset"] == "HFD"]
    fig7_iter_tables.append((
        f"Raw source data (y-axis, RQ2) — every HFD R² bootstrap iteration "
        f"per phase × feature ({len(rq2_hfd):,} rows).",
        _iter_table(rq2_hfd),
    ))

write_figure_sheet(wb, 7, [(fig7_header, fig7)] + fig7_iter_tables)
write_figure_sheet(wb, 8, [
    (f"Sub-table A — per-cell mean balanced accuracy "
     f"(source of the boxplot dots)", fig8_per_cell),
    (f"Sub-table B — per-group summary "
     f"(band vs ratio, slow vs gamma) per phase", fig8_group),
])


# --- Extra sheet: pooled feature data (input to every model) ---
def write_pooled_features(wb):
    if not (os.path.exists(CABLE1_FEAT_CSV) and os.path.exists(CABLE3_FEAT_CSV)):
        print("Skipping 'Pooled features' sheet: input CSVs not found.")
        return
    c1 = pd.read_csv(CABLE1_FEAT_CSV)
    c3 = pd.read_csv(CABLE3_FEAT_CSV)
    pooled = pd.concat([c1, c3], ignore_index=True)
    ws = wb.create_sheet("Pooled features (input)")
    ws.cell(row=1, column=1,
            value="Pooled feature table — Cable 1 + Cable 3 "
                  "(input source data for every model)").font = TITLE_FONT
    ws.cell(row=3, column=1,
            value=(f"{len(pooled)} recordings from "
                   f"{pooled['mouse'].nunique()} unique animals. "
                   "This is the exact 10a-derived feature matrix that fed all "
                   "bootstrap models. Every model average shown on the figures "
                   "was ultimately computed from these values.")).font = CAPTION_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=12)
    ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 50
    write_dataframe_at(ws, pooled, start_row=5, start_col=1)


# --- Extra sheet: RQ1 per-iteration bootstrap values ---
def write_rq1_iterations(wb):
    if not os.path.exists(RQ1_ITER_CSV):
        print(f"Skipping 'RQ1 iterations' sheet: {RQ1_ITER_CSV} not found "
              "(re-run 10b_bootstrap_full.py to generate it).")
        return
    df = pd.read_csv(RQ1_ITER_CSV)
    ws = wb.create_sheet("RQ1 iterations")
    ws.cell(row=1, column=1,
            value="RQ1 per-iteration bootstrap values "
                  "(numbers averaged into Figures 1, 2, 3, 7)").font = TITLE_FONT
    ws.cell(row=3, column=1,
            value=(f"{len(df):,} rows — one per (phase × feature × model × "
                   "bootstrap iteration). The 'balanced_accuracy' column holds "
                   "the raw values that were averaged into the mean and CI "
                   "shown on Figures 1, 2, 3, and 7. 84 (phase × feature) "
                   "cells × 2 models × 1000 iterations.")).font = CAPTION_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=12)
    ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 60
    write_dataframe_at(ws, df, start_row=5, start_col=1)


# --- Extra sheet: RQ2 per-iteration bootstrap values ---
def write_rq2_iterations(wb):
    if not os.path.exists(RQ2_ITER_CSV):
        print(f"Skipping 'RQ2 iterations' sheet: {RQ2_ITER_CSV} not found "
              "(re-run 11c1_bootstrap_weight_full.py to generate it).")
        return
    df = pd.read_csv(RQ2_ITER_CSV)
    ws = wb.create_sheet("RQ2 iterations")
    ws.cell(row=1, column=1,
            value="RQ2 per-iteration bootstrap values "
                  "(numbers averaged into Figures 4, 5, 7)").font = TITLE_FONT
    ws.cell(row=3, column=1,
            value=(f"{len(df):,} rows — one per (subset × phase × feature × "
                   "bootstrap iteration). The 'R2' column holds the raw R² "
                   "values that were averaged into the mean and CI shown on "
                   "Figures 4, 5, and 7. Target = absolute body weight; "
                   "subset ∈ {HFD, CTRL}.")).font = CAPTION_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=12)
    ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 60
    write_dataframe_at(ws, df, start_row=5, start_col=1)


write_pooled_features(wb)
# NOTE: per-figure iteration data is now embedded directly in each Figure
# sheet (as sub-tables below the summary), so we no longer emit standalone
# 'RQ1 iterations' / 'RQ2 iterations' sheets.

wb.save(OUT_XLSX)

print(f"\nSaved Excel workbook:\n{OUT_XLSX}")
print("Structure per figure sheet:")
print("  Row 1: title  |  Row 3: caption  |  Row 5+: embedded figure image")
print("  Below the image: source data table(s)")
